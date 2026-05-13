import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def checkGridTableAvailability(BannerFile):

    colNames = ['col' + str(colNo)  for colNo in range(0,23)]
        
    griddf = pd.read_excel(BannerFile,sheet_name='Grid',header=None,names=colNames)

    for col in griddf.columns:
        griddf[col] = griddf[col].astype('str')

    SummaryList = []
    for i in griddf['col1'].str.contains('Summary'):
        SummaryList.append(i)

    print(len(SummaryList))

    return len(SummaryList)
       
def getNoofColumns(BannerFile):

    griddf = pd.read_excel(BannerFile,sheet_name='Grid')

    return len(griddf.columns)

def getGridTableList(BannerFile):

    colNames = ['col' + str(colNo)  for colNo in range(0,getNoofColumns(BannerFile))]
        
    griddf = pd.read_excel(BannerFile,sheet_name='Grid',header=None,names=colNames)

    for col in griddf.columns:
        griddf[col] = griddf[col].astype('str')


    gridQuestions = []
    for questions in griddf[griddf['col0'].str.contains('Summary')].col0:
        gridQuestions.append(questions.split('.')[0])
    

    return list(set(gridQuestions))


def getSplitterInfo(Question):
     
    splitCount = []
    for ele in Question:
        if "[" in ele:
            splitCount.append(ele)
        
    return (len(splitCount))
     

def convertTablestoStandard(BannerFile):

    # if checkGridTableAvailability(BannerFile) > 0:
    #     print("Grid Tables are present the scripts are proceeding.\n")

        questionsList = getGridTableList(BannerFile)

        tables = pd.read_excel(BannerFile,sheet_name='Tables')

        for col in tables.columns:
            tables[col] = tables[col].astype('str')

        Grid_df = pd.DataFrame(
            {
                "Title": [] ,
                "Label": [] ,
                "TablesCount": [] ,
                "Percentile": []
            }
        )
        
        for question in  questionsList:
            question_index = tables.index[tables['Title'].str.contains(question)]

            start_index = question_index[4]
            end_index = question_index[-1]
            print(start_index)
            print(end_index)
            df = tables.iloc[start_index:end_index,:4]

            Grid_df = pd.concat([Grid_df,df])

        Grid_df['Question'] = Grid_df['Title'].apply(lambda x : x.split('.')[0])

        Grid_df = Grid_df.reset_index()

        Statements = []
        for title in Grid_df['Title']:
            if getSplitterInfo(title) == 1:
                Statements.append(title.split(' - ')[0].split('. ')[1].replace('[',"").replace("]",''))
            elif getSplitterInfo(title) > 1:
                Statements.append(title.split('] - [')[1].split('] - ')[0])
            else:
                Statements.append("")

        print(len(Statements))
        print(Grid_df.shape[0])

        Grid_df['Statements'] = Statements
        
        brandList = []
        for title in Grid_df['Title']:
            if getSplitterInfo(title) == 1:
                brandList.append('')
            elif getSplitterInfo(title) > 1:
                brandList.append(title.split('] - [')[0].split('. - [')[1])
            else:
                brandList.append('')

        print(len(brandList))
        Grid_df['Brands'] = brandList

        keyList = []
        for i in range(0,Grid_df.shape[0]):
            if getSplitterInfo(Grid_df['Title'][i]) == 1:
                keyList.append(Grid_df['Question'][i] + " - " + Grid_df['Statements'][i] + " - " + Grid_df['Label'][i])
            elif getSplitterInfo(Grid_df['Title'][i]) > 1:
                keyList.append(Grid_df['Question'][i] + " - " + Grid_df['Brands'][i] + " - " + Grid_df['Statements'][i] + " - " + Grid_df['Label'][i])
            else:
                keyList.append('')

        Grid_df['Key'] = keyList

        Grid_df.to_excel("Tables_Grid.xlsx",index=False)

        return Grid_df


def convertGridstoStandard(BannerFile):

    NoOfColumns = getNoofColumns(BannerFile)
    print(f'Added Banner has {NoOfColumns}\n')
    colNames = ['col' + str(colNo)  for colNo in range(0,NoOfColumns - 1)]

    Grid = pd.read_excel(BannerFile,sheet_name='Grid',header=None,names=colNames)
    Grid.dropna(how= 'all',inplace=True)
    Grid = Grid.reset_index()

    for col in Grid.columns:
        Grid[col] = Grid[col].astype('str')

    start_index = Grid.index[Grid['col0'].str.contains('Table')] - 1
    end_index = Grid.index[Grid['col0'].str.contains('Grid overlap formula')] 

    indices = list(zip(start_index,end_index))

    Title = []
    Statements = []
    Label = []
    Counts = []


    for indexes in indices:
        df = Grid.iloc[indexes[0]:indexes[1]]

        Statement_index = indexes[0] + 2

        TableLength = abs(Statement_index - indexes[1])

        print(f'The table Lenght is {TableLength}')

        for colNo in range(1,NoOfColumns - 1):
            for i in range(0,TableLength):
                Title.append(df['col0'][indexes[0]])
                Statements.append(df['col' + str(colNo)][Statement_index])
                

        for colNo in range(1,NoOfColumns - 1):
                for i in (df.loc[Statement_index:indexes[1]]['col0']):
                     Label.append(i)
                for k in (df.loc[Statement_index:indexes[1]]['col' + str(colNo)]):
                     Counts.append(k)



    Grid_df = pd.DataFrame(
         {
              'Title' : Title,
              'Statements' : Statements,
              'Label' : Label,
              'GridsCounts' : Counts 
         }
    )

    Grid_df['Question'] = Grid_df['Title'].apply(lambda x : x.split('.')[0])

    brandList = []
    for title in Grid_df["Title"]:
        if getSplitterInfo(title) == 1:
            brandList.append('')
        elif getSplitterInfo(title) > 1:
            brandList.append(title.split('] - [')[0].split('. - [')[1])
        else:
            brandList.append('')

    Grid_df['Brands'] = brandList
    

    
    keyList = []
    for i in range(0,Grid_df.shape[0]):
        if getSplitterInfo(Grid_df['Title'][i]) == 1:
            keyList.append(Grid_df['Question'][i] + " - " + Grid_df['Statements'][i] + " - " + Grid_df['Label'][i])
        elif getSplitterInfo(Grid_df['Title'][i]) > 1:
            keyList.append(Grid_df['Question'][i] + " - " + Grid_df['Brands'][i] + " - " + Grid_df['Statements'][i] + " - " + Grid_df['Label'][i])
        else:
            keyList.append('')


    Grid_df['Key'] = keyList

    Grid_df.to_excel('Grid_tables.xlsx',index=False)

    return Grid_df


def combiningAndComparingGridTables(BannerFile,OuputDir):

    if checkGridTableAvailability(BannerFile) > 0:

        GridTables = convertGridstoStandard(BannerFile)
        Tables = convertTablestoStandard(BannerFile)


        Combined_df = pd.merge(Tables,GridTables,on = 'Key',how='inner')

        Cleaned_Combined_df = Combined_df[['Title_x','Statements_x','Label_x','GridsCounts','Count','Key']]

        Result = []
        for cell in range(0,Cleaned_Combined_df.shape[0]):
            if Cleaned_Combined_df['GridsCounts'][cell] == Cleaned_Combined_df['Count'][cell]:
                Result.append('True')
            if Cleaned_Combined_df['GridsCounts'][cell] != Cleaned_Combined_df['Count'][cell]:
                Result.append('False')

        Cleaned_Combined_df['Result'] = Result
            
        FileDumpDir = os.path.join(OuputDir,"Final Comparison.xlsx")
        
        print(f"The final comparison file is dumped at the location {FileDumpDir}\n")
        wb = load_workbook(FileDumpDir)
        
        light_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        light_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        if "Grid Comparison" in wb.sheetnames:
            wb.remove(wb["Grid Comparison"])
        
        ws = wb.create_sheet("Grid Comparison")

        for r in dataframe_to_rows(Cleaned_Combined_df, index=False, header=True):
            ws.append(r)


        for row in ws.iter_rows():   
            for cell in row:
                if cell.value == "True":
                        cell.fill = light_green 
                elif cell.value == "False":
                        cell.fill = light_red    

        wb.save(FileDumpDir)

# Banner = r'C:\Users\Irshad.kazi\OneDrive - Ipsos\Desktop\Secondary QC Automation\_Versions_\QC - Automation V9\Input\Banners.xlsx'
# OutputDir = r'C:\Users\Irshad.kazi\OneDrive - Ipsos\Desktop\Secondary QC Automation\_Versions_\QC - Automation V9\Output'
# combiningAndComparingGridTables(Banner,OutputDir)

# convertGridstoStandard(Banner)
# convertTablestoStandard(Banner)