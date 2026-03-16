import pandas as pd 
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from openpyxl.styles import Font


def loadCountFile(path):
    try:
        columnNames = []
        for i in range(0,5):
            columnNames.append('col' + str(i))

        df = pd.read_excel(path,header=None,names=columnNames)
        return df
    except Exception as e:
        print(f'Error While Loading the File at path\n {path}\n error : {e}')


def GenerateGridVariableDataset(path):
    try :
        df = loadCountFile(path)

        for col in df.columns:
            df[col] = df[col].astype('str')
            

        variableList = list(df['col1'].to_list())

        gridVariables = [var.split(' VARIABLE_LABEL')[0]  for var in variableList if "|" in var]

        stubsInfo = []
        CountsInfo = []
        for gridVar in gridVariables:

            start_ind = list(df.index[df['col0'].str.contains(gridVar,regex=False)])[0] + 3
            end_ind = list(df.index[df['col0'].str.contains(gridVar,regex=False)])[-1]

            Stage1df = df.loc[start_ind:end_ind,['col1','col2']].reset_index()

            variableNameLabelBrand = Stage1df.loc[0,'col1']
            variableName = variableNameLabelBrand.split(' VARIABLE_LABEL')[0].split(' = ')[1]
            variableName = re.sub(r"\[\{_?[A-Za-z0-9]+\}\]", '', variableName)
            variableNameSegment = variableName.split('[')[0]
            variableBrand = variableNameLabelBrand.split(' || ')[1]
            Stage1df.iat[0,1] = "Segment : "  + variableName
            Stage1df.iat[0,2] = variableBrand
            
            for i in Stage1df['col1']:
                if i == 'nan':
                    stubsInfo.append('')
                else:
                    stubsInfo.append(i)
            
            for k in Stage1df['col2']:
                if k == 'nan':
                    CountsInfo.append('')
                else:
                    CountsInfo.append(k)

        GridVariabledf = (pd.DataFrame({"GridsStubs":stubsInfo,'Counts':CountsInfo}))
        GridVariabledf.to_excel('GridVarDataset.xlsx',index=False)
        return GridVariabledf
    except Exception as e:
        print(f'Error while creating Grid variables dataFrame : {e}')


def GenerateGridTables(file_path, output_path):

    df = GenerateGridVariableDataset(file_path)
    
    table_start = list(df.index[df['GridsStubs'].str.contains('Segment',regex=False)])
    table_end = list(df.index[df['GridsStubs'].str.contains(';')])
    
    # QuestionNames = [df.loc[i]['GridsStubs'].split(' : ')[1] for i in df.index[df['GridsStubs'].str.contains('Segment')]]
    QuestionNames = [df.loc[i]['GridsStubs'] for i in df.index[df['GridsStubs'].str.contains('Segment',regex=False)]]

    print(f'No. of Questions : {len(QuestionNames)}')
    print(f'No. of Start Points : {len(table_start)}')
    print(f'No. of End Points : {len(table_end)}')
    
    if len(QuestionNames) == len(table_start) == len(table_end):
        TableInfodf = (pd.DataFrame({'Questions':QuestionNames,'StartPoints':table_start,'EndPoints':table_end}))

        QuestionList = TableInfodf['Questions'].unique()

        wb = openpyxl.load_workbook(output_path)
        activeSheets = wb.sheetnames
        print(activeSheets)
        if 'Grid Tables' in activeSheets:
            del wb['Grid Tables']
                
        wb.create_sheet('Grid Tables')

        ws = wb['Grid Tables']

        start_row = 1
        for j,questions in enumerate(QuestionList):
            df1 = TableInfodf[TableInfodf['Questions'].str.contains(questions,regex=False)].reset_index()
            # print(df1)
            startpt = df1['StartPoints'][0]
            endpt = df1.loc[df1.shape[0] - 1]['EndPoints']


            if j == 0:
                start_row = 1
                tableSize = abs(df1['StartPoints'][0] - df1['EndPoints'][0]) + 5
            else:
                start_row = start_row + tableSize
                tableSize = abs(df1['StartPoints'][0] - df1['EndPoints'][0]) + 5
                
            
            print(f'Start Row for Question {questions} : {start_row}')
            print(f'table Size for Question {questions} : {tableSize}')

            for i in range(0,df1.shape[0] - 1):
                if i == 0:
                    start_col = 1
                    for r_idx, row in enumerate(dataframe_to_rows(df.loc[df1.loc[i]['StartPoints'] : df1.loc[i]["EndPoints"], ['GridsStubs','Counts']], index=False, header=False), start_row):
                        for c_idx, value in enumerate(row, start_col):
                            ws.cell(row=r_idx, column=c_idx, value=value)

                    start_col = 3
                elif i > 0:
                    for r_idx, row in enumerate(dataframe_to_rows(df.loc[df1.loc[i]['StartPoints'] : df1.loc[i]["EndPoints"],['Counts']], index=False, header=False), start_row):
                        for c_idx, value in enumerate(row, start_col):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                
                            wb.save(output_path)

                    start_col = start_col + 1     
            print(f'Completed Question : {questions}\n')

        for cell in ws['A']:
            cell.font = Font(bold=True,color="FF00008B")

        wb.save(output_path)