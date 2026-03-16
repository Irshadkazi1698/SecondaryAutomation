import pandas as pd 
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# PATH = r'C:\Users\Irshad.kazi\OneDrive - Ipsos\Desktop\Secondary QC Automation\_Versions_\QC - Automation V5\Output\Final Comparison.xlsx'
# TabPlanPath = r'C:\Users\Irshad.kazi\OneDrive - Ipsos\Desktop\Secondary QC Automation\_Versions_\QC - Automation V5\Input\MoodBoardTBPlan_V6.xlsm'


def SigmaCheck(path):
    df = pd.read_excel(path,sheet_name='Tables')

    for i in df.columns:
        df[i] = df[i].astype('str')
    
    Sigma_df = df[df['Label'].str.contains('Sigma')][['Title','Label','Percentile']]

    Sigma_df['Percentile'] = Sigma_df['Percentile'].apply(lambda x : int(float(x)*100))

    title = []
    percent = []
    for i in range(0,Sigma_df.shape[0]):
        if int(Sigma_df.iloc[i,2]) < 100:
            title.append(Sigma_df.iloc[i,0])
            percent.append(int(Sigma_df.iloc[i,2])*100)

    Sigma_Final_df = pd.DataFrame({'TableName':title,'SigmaValue':percent})


    if Sigma_Final_df.shape[0] >= 1:
        return Sigma_Final_df
    else:
        return 'Sigma values are aligned!'


def TitleComparisonCheck(path):

    df = pd.read_excel(path,sheet_name='Tables')

    for col in df.columns:
        df[col] = df[col].astype('str')

    TitleIndex = df.index[df['Label'].str.contains('Table:')] - 1

    df = df.iloc[TitleIndex, : ]

    df = df[['Label','Label.1']]


    NotMatchedTitle = []
    for i in range(0,df.shape[0]):
        if df.iloc[i]['Label'] != df.iloc[i]['Label.1']:
            NotMatchedTitle.append(df.iloc[i]['Label'])

    df = pd.DataFrame({'TitleNotMatched':NotMatchedTitle})

    if df.shape[0] >= 1:
        return df
    else:
        return 'Not mismatched Title'
    

def BaseTextComparisonCheck(path):

    df = pd.read_excel(path,sheet_name='Tables')

    for col in df.columns:
        df[col] = df[col].astype('str')

    BaseIndex = df.index[df['Label'].str.contains('Table:')] + 1

    df = df.iloc[BaseIndex, : ]

    df = df[['Label','Label.1']]

    NotMatchedBaseText = []
    for i in range(0,df.shape[0]):
        if df.iloc[i]['Label'] == df.iloc[i]['Label.1']:
            NotMatchedBaseText.append(df.iloc[i]["Label"])


    df = pd.DataFrame({'BaseTextNotMatched':NotMatchedBaseText})

    if df.shape[0] >= 1:
        return df
    else:
        return 'No Mismatch in the base text.'

   
def VariableCountsCheck(path,tabplanpath):

    dftp = pd.read_excel(tabplanpath,sheet_name ='STUB SPECS',skiprows=1)

    for col in dftp.columns:
        dftp[col] = dftp[col].astype('str')

    tabplanQuestion = sorted(set(dftp['Question #'].apply(lambda x : x.split('[')[0]).to_list()))


    df = pd.read_excel(path)

    for col in df.columns:
        df[col] = df[col].astype('str')
    
    TableQuestion = sorted(set(df['Title'].apply(lambda x : x.split('.')[0]).to_list()))

    MissingTables = []
    for i in tabplanQuestion:
        if i not in TableQuestion:
            MissingTables.append(i)

    df = pd.DataFrame({'MissingTables': MissingTables})

    if df.shape[0] >= 1:
        return df
    else:
        return 'No Missing tables.'
    
def checkBaseSize(path):

    df = pd.read_excel(path,sheet_name='Tables')

    for col in df.columns:
        df[col] = df[col].astype('str')
    
    baseSizeIndex = df.index[df['Label'].str.contains('Table:')] + 1

    df = df.iloc[baseSizeIndex][['Label','Count','Label.1','Count.1']]

    df = df.groupby('Label')['Count'].value_counts().reset_index()

    df = df['Label'].value_counts().reset_index()

    df = df[df['count'] >1]

    if df.shape[0] >= 1:
        return df
    else:
        return 'No Duplicated Base Sizes.'
    

def getJunkCharacter(path):

    df = pd.read_excel(path,sheet_name='Tables')

    junkchar = ["@","#"]

    for col in df.columns:
        df[col] = df[col].astype('str')

    JunkLabel = []
    for char in junkchar:
        junkcharind = df.index[df['Label'].str.contains(char)]
        for i in junkcharind:
            JunkLabel.append(df.iloc[i]['Label'])
        
    
    JunkLabel = list(set(JunkLabel))

    df = pd.DataFrame({'VarContainJunkValues':JunkLabel})

    if df.shape[0] >= 1:
        return df
    else:
        return "No Junk characters."



def createSanityCheck(path,tabplanpath):

    wb = openpyxl.load_workbook(path)

    sheets = wb.sheetnames
    if 'Sanity Check' in sheets:
        del wb['Sanity Check']
    
    wb.create_sheet(title = 'Sanity Check')

    navy_blue_hex = "000080"
    thick_border = Side(border_style="thick", color="000000")
    full_border = Border(left=thick_border, right=thick_border, top=thick_border, bottom=thick_border)

    ws = wb['Sanity Check']
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['B'].width = 15

    '''
    Sigma Check - Starts here.
    '''
    ws['A1'] = 'Sigma Check'
    ws['A1'].font = Font(bold=True,size=18,color=navy_blue_hex)

    SigmaData =  SigmaCheck(path)

    green_hex = "FF008000"

    if type(SigmaData) is str:
        ws['A2'] = SigmaData
        ws['A2'].font = Font(color=green_hex)
    else:
        for row in dataframe_to_rows(SigmaData,index=False,header=True):
            ws.append(row)

    ws.cell(row=2,column=2).font = Font(bold=True)
    ws.cell(row=2,column=1).font = Font(bold=True)

    print("Sigma Check Completed!")
    '''
    Sigma Check - End here.
    '''

    '''
    Title Check - Start here.
    '''
    if type(SigmaData) == str:
        PaddingSize = 5
    else:
        DataSize = SigmaData.shape[0]
        PaddingSize = DataSize + 5

    TitleStartCell = 'A' + str(PaddingSize)

    ws[TitleStartCell] = 'Title Check'
    ws[TitleStartCell].font = Font(bold=True,size=18,color=navy_blue_hex)

    titleData = TitleComparisonCheck(path)

    if type(titleData) is str:
        titleResultCell = 'A' + str(PaddingSize + 1)
        ws[titleResultCell] = titleData
        ws[titleResultCell].font = Font(bold=True)
    else: 
        for row in dataframe_to_rows(titleData):
            ws.append(row)

    ws['B' + str(PaddingSize + 2)].font = Font(bold=True)

    print("Title Check Completed!")
    
    '''
    Title Check - End here
    '''

    '''
    Base Text Check - Start here
    '''
    if type(titleData) is str:
        PaddingSize = PaddingSize + 7
    else :
        PaddingSize = PaddingSize + 1 + titleData.shape[0] + 5
    
    ws["A" + str(PaddingSize)] = 'Base Text Check'
    ws["A" + str(PaddingSize)].font = Font(bold=True,size=18,color=navy_blue_hex)

    BaseData = BaseTextComparisonCheck(path)
    
    if type(BaseData) is str:
        ws["A" + str(PaddingSize + 2)] = BaseData
        ws["A" + str(PaddingSize + 2)].font = Font(bold=True)
    else:
        for rows in dataframe_to_rows(BaseData):
            ws.append(rows)

    print("Base Text Check Completed!")
    '''
    Base Text Check - End here
    '''

    '''
    Missing Table - Start here
    '''
    if type(BaseData) is str:
        PaddingSize = PaddingSize + 2 + 5
    else:
        PaddingSize = PaddingSize + 1 + BaseData.shape[0] + 5

    ws['A' + str(PaddingSize)] = 'Missing Tables'
    ws['A' + str(PaddingSize)].font = Font(bold = True, size = 18, color = navy_blue_hex )

    Missingtabledata = VariableCountsCheck(path,tabplanpath)

    if type(Missingtabledata) is str:
        ws["A" + str(PaddingSize + 2)] = Missingtabledata
        ws["A" + str(PaddingSize + 2)].font = Font(bold = True,color=green_hex)
    else:
        for row in dataframe_to_rows(Missingtabledata):
            ws.append(row)        

    print("Missing Table Check Completed!")

    '''
    Missing Table - End here
    '''

    '''
    Base Size Check - Start here
    '''
    if type(Missingtabledata) is str:
        PaddingSize = PaddingSize + 2 + 5
    else:
        PaddingSize = PaddingSize + 1 + Missingtabledata.shape[0] + 5

    ws["A" + str(PaddingSize)] = "Base Size Check"
    ws["A" + str(PaddingSize)].font = Font(bold=True,size=18,color=navy_blue_hex)

    BaseSizeData = checkBaseSize(path)

    if type(BaseSizeData) is str:
        ws["A" + str(PaddingSize + 1)] = BaseSizeData
        ws["A" + str(PaddingSize + 1)].font = Font(bold=True,color = green_hex)
    else:
        for row in dataframe_to_rows(BaseSizeData):
            ws.append(row)

    print("Base Size Check Completed!")
    '''
    Base Size Check - End here
    '''
    
    '''
    Junk Character Check - Start here
    '''
    
    if type(BaseSizeData) is str:
        PaddingSize = PaddingSize + 2 + 5
    else:
        PaddingSize = PaddingSize + 1 + BaseSizeData.shape[0] + 5
    
    ws["A" + str(PaddingSize)] = "Junk Character Check"
    ws["A" + str(PaddingSize)].font = Font(bold=True,size=18,color=navy_blue_hex)

    JunkCharData = getJunkCharacter(path)

    if type(JunkCharData) is str:
        ws["A" + str(PaddingSize + 2)] = JunkCharData
        ws["A" + str(PaddingSize + 2)].font = Font(bold=True,color = green_hex)
    else:
        for row in dataframe_to_rows(JunkCharData):
            ws.append(row)

    print("Junk Character Check Completed!")
    '''
    Junk Character Check - End here 
    '''
    
    
    wb.save(path)
