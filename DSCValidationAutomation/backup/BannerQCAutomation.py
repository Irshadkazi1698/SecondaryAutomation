import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz, process
import re
import os
import sys
from collections import defaultdict
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

# --- CONFIGURATION & CONSTANTS ---

# # These are the names we'll be using for our output Excel files.
# COUNT_FILE_NAME = 'Counts'
# BANNER_FILE_NAME = 'Banners'
# UPDATED_MATCHING_FILE_NAME = 'Matched_Variables'
# SUMMARY_FILE_NAME = 'Unmatched_Summary'
# FINAL_COMPARISON_FILE_NAME = 'Final Comparison'

# # Here's where the script expects to find your input files and where it will save the results.
# INPUT_FOLDER_PATH = r'C:\Users\Irshad.kazi\OneDrive - Ipsos\Desktop\Secondary QC Automation\QC - Automation\Input/'
# OUTPUT_FOLDER_PATH = r'C:\Users\Irshad.kazi\OneDrive - Ipsos\Desktop\Secondary QC Automation\QC - Automation\Output/'

# # This is a specific marker the script looks for in the 'Counts' file to identify variable names.
# VARIABLE_PREFIX = 'VARIABLE_NAME = '

# --- HELPER & PROCESSING FUNCTIONS ---

def generate_unmatched_summary(matching_df, output_path):
    """
    This function takes the matching results and creates a text file that summarizes which table titles
    couldn't be automatically matched. It also provides a quick overview of how many tables were matched.
    """
    print("Generating a summary of unmatched table titles...")
    # First, we make sure the 'Definition' column is clean and treated as text, filling any empty spots.
    matching_df['Definition'] = matching_df['Definition'].fillna('').astype(str)
    
    # We define a pattern to identify and exclude certain types of summary rows that aren't relevant for matching.
    exclude_pattern = r'\[Top 2 Box - Summary\]|\[Bottom 2 Box - Summary\]|\[Summary - Mean\]'
    # Then, we filter out those specific summary rows from our DataFrame.
    filtered_matching_df = matching_df[~matching_df['Definition'].str.contains(exclude_pattern, regex=True)]

    # Now, let's calculate some statistics for the summary.
    total_entries = len(filtered_matching_df) # Total number of tables we're considering.
    filled_entries = filtered_matching_df['Matched_Label'].notnull().sum() # How many successfully matched.
    blank_entries = total_entries - filled_entries # How many were left unmatched.
    percentage_matched = (filled_entries / total_entries) * 100 if total_entries > 0 else 0 # Percentage matched.

    # We'll write these findings into a .txt file.
    with open(output_path, 'w') as f:
        f.write("Unmatched Titles Summary\n========================\n\n")
        # List out all the table titles that didn't find a match.
        unmatched_titles = filtered_matching_df[filtered_matching_df['Matched_Label'].isnull()]['Title']
        for title in unmatched_titles:
            f.write(f"{title}\n") # Write each unmatched title on a new line.
        
        # Now, add the summary statistics to the file.
        f.write("\nSummary Statistics\n==================\n")
        f.write(f"Total Entries: {total_entries}\nFilled Entries: {filled_entries}\n")
        f.write(f"Blank Entries: {blank_entries}\nPercentage of Tables Matched: {percentage_matched:.2f}%\n")
    print(f"Summary of unmatched tables has been successfully written to: {output_path}")

def create_and_prepare_output_file(source_file_path, output_file_path, sheet_name='Tables'):
    """
    This function sets up our main output Excel file. If it doesn't exist, it creates it.
    Then, it copies the structure, data, and formatting from the 'Banners' file (which serves as our template)
    into the output file. This ensures our final output looks just like the original banner layout.
    """
    print(f"Getting the output file ready at: '{output_file_path}'...")
    
    # First, let's check if the output file already exists. If not, we create a new one.
    if not os.path.exists(output_file_path):
        wb = Workbook() # Create a new workbook.
        ws = wb.active # Get the active sheet.
        ws.title = sheet_name # Name it 'Tables'.
        ws.sheet_view.showGridLines = True # Make sure the grid lines are visible.
        wb.save(output_file_path) # Save the newly created file.

    # Now, we open both the output file (where we'll write) and the source file (for copying)
    # We load the source file twice: once to get the actual data values, and once to get the styles.
    dest_wb = load_workbook(output_file_path) # Open the destination workbook.
    dest_ws = dest_wb[sheet_name] # Get the 'Tables' sheet from the destination.
    source_wb = load_workbook(source_file_path, data_only=True) # Load source for data values.
    source_ws = source_wb[sheet_name] # Get the 'Tables' sheet from the source for data.

    # Load the source workbook again, but this time without `data_only=True`, so we can grab styles.
    source_wb_for_styles = load_workbook(source_file_path, data_only=False)
    source_ws_for_styles = source_wb_for_styles[sheet_name]

    # We'll go through each column and copy its width from the source.
    max_col = source_ws.max_column # Find out how many columns we have.
    for i in range(1, max_col + 1):
        col_letter = get_column_letter(i) # Get the letter for the current column (e.g., 'A', 'B').
        # Check if the source column has a specific width set.
        if source_ws_for_styles.column_dimensions[col_letter].width:
             # If yes, apply that width to our destination column.
             dest_ws.column_dimensions[col_letter].width = source_ws_for_styles.column_dimensions[col_letter].width
        else: # If no specific width was set in the source, use a default.
            dest_ws.column_dimensions[col_letter].width = 8.43

    # Next, we copy each row, including its height and all cell contents and styles.
    max_row = source_ws.max_row # Find out how many rows we have.
    for r in range(1, max_row + 1):
        # First, copy the row height.
        if source_ws_for_styles.row_dimensions[r].height:
            dest_ws.row_dimensions[r].height = source_ws_for_styles.row_dimensions[r].height

        # Now, iterate through each cell in the current row.
        for c in range(1, max_col + 1):
            source_cell_val = source_ws.cell(row=r, column=c) # Get the cell with its data value.
            source_cell_style = source_ws_for_styles.cell(row=r, column=c) # Get the cell with its styling.
            target_cell = dest_ws.cell(row=r, column=c) # Get the corresponding cell in our output file.
            
            target_cell.value = source_cell_val.value # Copy the data value from the source.

            # If the source cell has any styles applied, copy them over.
            if source_cell_style.has_style:
                target_cell.font = copy(source_cell_style.font) # Copy font.
                target_cell.alignment = copy(source_cell_style.alignment) # Copy alignment.
                target_cell.fill = copy(source_cell_style.fill) # Copy fill color.
                target_cell.border = copy(source_cell_style.border) # Copy borders.
                target_cell.number_format = source_cell_style.number_format # Copy number format.
                target_cell.protection = copy(source_cell_style.protection) # Copy protection settings.
    
    # Finally, save all the changes to our output file.
    dest_wb.save(output_file_path)
    print("Output file created and copied formatting successfully.")
    
def get_variable_names_from_counts(counts_df, prefix):
    """
    This function is designed to pull out all the unique variable names from the 'Counts' Excel file.
    It looks for a specific prefix that marks the beginning of a variable name.
    """
    variable_names = set() # We use a set to automatically handle duplicates.
    # We'll go through each entry in the first column of the counts DataFrame.
    for row in counts_df.iloc[:, 0]:
        # We only care about rows that are strings and contain our special prefix.
        if isinstance(row, str) and prefix in row:
            try:
                # If the prefix is found, we split the string to get the variable name part.
                # We take everything after the prefix and then split by space to isolate the variable name itself.
                variable_name = row.split(prefix)[1].split(' ')[0]
                variable_names.add(variable_name) # Add the found variable name to our set.
                
                # Sometimes, 'DP VARIABLE' indicates a slightly different naming convention,
                # so we also check for that and add the underscore-separated version if found.
                if "DP VARIABLE" in row:
                    variable_name_underscore = row.split(prefix)[1].split('_')[0]
                    if variable_name_underscore: # Make sure we actually got a name.
                        variable_names.add(variable_name_underscore)
            except IndexError:
                # If there's an issue splitting the string (e.g., prefix is at the very end), we just skip that row.
                continue
    # Once we've processed all rows, we return the unique variable names as a list.
    return list(variable_names)

def fetch_tables_indices(formatting_df, prefix, tab_title, mode):
    """
    This function is crucial for locating specific tables within the data. It finds the starting
    and ending rows for a given table title, based on whether we're looking in the 'Count' data
    or the 'Banner' data. It uses specific patterns to identify the table boundaries.
    """
    range_table_indices = [] # This list will store the (start_row, end_row) for each found table.
    
    if mode == 'Count':
        # For 'Count' mode, we use a pattern that's designed to catch the table titles followed by a space.
        # This pattern is taken directly from your working script and is meant to be robust.
        pattern = rf"{prefix}{re.escape(tab_title.upper())}\s"
        
        # We find all rows where the second column (converted to uppercase) matches our pattern.
        start_indices = formatting_df[formatting_df.iloc[:, 1].str.upper().str.contains(pattern, regex=True, na=False)].index.tolist()
        
        # We also need to find rows that indicate the start of a *new* table or a break.
        # The presence of a semicolon (';') in the second column often signifies this.
        overlap_indices = formatting_df[formatting_df.iloc[:, 1].str.upper().str.contains(";", regex=True, na=False)].index.values.tolist()
        
        # Now, for each identified start index, we find the next "overlap" index to determine the end of the table.
        for idx in start_indices:
            # We look for overlap indices that are *after* the current start index.
            larger = [x for x in overlap_indices if x > idx]
            # The end of the table is either the row *before* the next overlap index, or the very end of the DataFrame if no more overlaps occur.
            range_table_indices.append((idx, min(larger) - 1 if larger else len(formatting_df)))
            
    elif mode == 'Banner':
        # For 'Banner' mode, the pattern is simpler: it just looks for the exact table title at the end of the second column.
        pattern = f"{re.escape(tab_title)}$"
        start_indices = formatting_df[formatting_df.iloc[:, 1].str.contains(pattern, regex=True, na=False)].index.tolist()
        
        # Similar to 'Count' mode, we look for rows indicating breaks, but this time using '<BR/>' as a delimiter.
        overlap_indices = formatting_df[formatting_df.iloc[:, 1].str.contains(" <BR/> ", regex=False, na=False)].index.values.tolist()
        
        # We determine the end of the table based on these overlap indices.
        for idx in start_indices:
            larger = [x for x in overlap_indices if x > idx]
            # The end is the next overlap index, or the end of the DataFrame if none exist.
            # Note: Unlike 'Count' mode, we take the overlap index itself as the end boundary here.
            range_table_indices.append((idx, min(larger) if larger else len(formatting_df)))
            
    return range_table_indices

def align_counts_with_banners(banners_formatting, counts_formatting, key_col, base_values_list, threshold=100):
    """
    This is a core function where we try to match rows from the 'Counts' data to the 'Banners' data.
    It's smart enough to handle exact matches, fuzzy matches (using fuzzywuzzy), and special 'BASE'
    labels, aligning them with corresponding data from the 'Counts' file.
    """
    # Ensure the key columns are treated as strings for consistent matching.
    banners_formatting[key_col] = banners_formatting[key_col].astype(str)
    counts_formatting[key_col] = counts_formatting[key_col].astype(str)
    
    aligned_counts = [] # This will store the matched rows from the counts data.
    # Create a quick lookup dictionary from the counts DataFrame for efficient searching.
    counts_dict = {row[key_col]: row for _, row in counts_formatting.iterrows()}
    available_choices = counts_formatting[key_col].unique() # Get all unique labels from counts for fuzzy matching.

    # A helper function to find the best fuzzy match for a given label.
    def find_best_match(label, choices):
        # This part cleans up labels before matching, like removing "(score)" text.
        if '(' in label and 'Net' not in label and re.search(r'\(\d+(\.\d+)?\)', label):
            label = label.split('(')[0].strip()
            
        # Use fuzzywuzzy's extractOne to find the best match and its score.
        best_match, highest_score = process.extractOne(label, choices, scorer=fuzz.ratio)
        # Only return the match if the score meets our threshold (default is 100 for exact match).
        return best_match if highest_score >= threshold else None

    # Helper function to specifically find the 'BASE' row in the counts data.
    def find_base_count_data():
        """Searches the counts DataFrame for a row labeled 'BASE' and returns it."""
        for _, row in counts_formatting.iterrows():
            if row[key_col].upper() == 'BASE': # Case-insensitive check for 'BASE'.
                return row # Return the entire row if found.
        return None # Return nothing if no 'BASE' row is found.

    # Now, we iterate through each row in the processed banners data.
    for _, banner_row in banners_formatting.iterrows():
        banner_label = banner_row[key_col] # Get the label from the current banner row.

        # Special handling for common 'Net' or 'Subnet' labels: just add an empty row.
        if banner_label in ['Box (Net)', 'Box (Subnet)', 'Net', '(Net)']:
            aligned_counts.append(pd.Series([np.nan] * len(counts_formatting.columns), index=counts_formatting.columns))
            
        # --- THIS IS THE NEW LOGIC FOR HANDLING 'BASE' LABELS ---
        # If the banner label starts with 'base' or matches a specific pattern indicating base respondents.
        elif banner_label.lower().startswith('base') or banner_label.startswith('Base: Respondents that answer >0 in Q.KIDS02'):
            if base_values_list: # Check if we still have base text values available.
                # Take the next available base text from our list.
                base_text_from_list = base_values_list.pop(0)
                # Construct the desired label format.
                new_label_value = f"Base: {base_text_from_list}"
                
                # *** NOW, FIND THE ACTUAL COUNT DATA FOR THE BASE ROW ***
                base_count_row = find_base_count_data()
                
                if base_count_row is not None:
                    # If we found the BASE row in counts, copy all its data.
                    new_row = base_count_row.copy()
                    # Update only the label column with our specially formatted base text.
                    new_row[key_col] = new_label_value
                    aligned_counts.append(new_row) # Add this modified row to our results.
                else:
                    # If no BASE row was found in the counts, create a row with just the label and NaNs for other data.
                    new_row = pd.Series([new_label_value] + [np.nan] * (len(counts_formatting.columns) - 1), index=counts_formatting.columns)
                    aligned_counts.append(new_row)
            else:
                # If we've run out of base text values, we'll just add an empty row as a fallback and print a warning.
                print("Warning: Ran out of values in 'Base Text' list while processing a base label.")
                aligned_counts.append(pd.Series([np.nan] * len(counts_formatting.columns), index=counts_formatting.columns))
                
        # For all other, non-base labels, we proceed with the standard fuzzy matching.
        else:
            best_match = find_best_match(banner_label, available_choices) # Find the best match in counts.
            if best_match:
                # If a match is found, append the corresponding row from the counts dictionary.
                aligned_counts.append(counts_dict[best_match])
            else:
                # If no match is found, append an empty row.
                aligned_counts.append(pd.Series([np.nan] * len(counts_formatting.columns), index=counts_formatting.columns))

    # Finally, convert the list of aligned rows into a DataFrame. If the list is empty, create an empty DataFrame with the correct columns.
    return pd.DataFrame(aligned_counts) if aligned_counts else pd.DataFrame(columns=counts_formatting.columns)

def populate_comparison_sheet(file_path, counts_df, banners_df_processed, banners_df_original, common_indices, base_texts_map,variablePrefix):
    """
    This function is responsible for actually pasting the data into the final output Excel file.
    It takes the aligned 'Counts' data and the original 'Banners' data and places them side-by-side
    in the correct locations within the output sheet.
    """
    print("Aligning and populating the 'Tables' sheet with data...")
    # Load the output workbook to start writing data.
    workbook = load_workbook(file_path)
    sheet = workbook['Tables']
    percentage_format = '0.00%' # Define how percentages should be displayed.

    # We need a way to quickly find which 'Matched_Label' (from the matching file) corresponds to which block of data
    # in the 'Counts' file. This map helps us retrieve the correct 'Base Text' values later.
    count_idx_to_label_map = {}
    for i, (fetch_idx, target_idx) in enumerate(common_indices):
        if fetch_idx and target_idx:
            # Get the very first label from the 'Counts' data for this table to identify the variable.
            first_label_in_count_table = counts_df.iloc[fetch_idx[0][0]]['Label']
            if isinstance(first_label_in_count_table, str) and first_label_in_count_table.startswith(variablePrefix):
                # Extract the base variable name part.
                variable_part = first_label_in_count_table.replace(variablePrefix, "").split(' ')[0]
                
                # Use fuzzy matching to confidently link this variable name to its 'Matched_Label' from our map.
                # We use a high threshold (97) to ensure accuracy.
                matched_label, score = process.extractOne(variable_part, base_texts_map.keys())
                if score >= 97:
                    # Store the starting row index of the count table and its corresponding Matched_Label.
                    count_idx_to_label_map[fetch_idx[0][0]] = matched_label

    # Now, we iterate through each identified table block to place its data.
    for i, (fetch_idx, target_idx) in enumerate(common_indices):
        if not fetch_idx or not target_idx: continue # Skip if we don't have valid indices for both count and banner.
        try:
            # --- IMPORTANT: Get BASE TEXTS SPECIFIC TO THIS TABLE ---
            # We look up the 'Matched_Label' associated with the current count table's starting index.
            current_matched_label = count_idx_to_label_map.get(fetch_idx[0][0])
            
            # Then, we retrieve the list of 'Base Text' values for that specific label.
            # We use .copy() so that when `align_counts_with_banners` uses .pop(), it doesn't affect other tables.
            specific_base_texts_for_this_table = base_texts_map.get(current_matched_label, []).copy()

            # --- Extracting Data for Logic and Writing ---
            # For the fuzzy matching logic, we need a slice of the *processed* banners DataFrame.
            # We subtract 2 from target_idx[0][0] because openpyxl is 1-indexed and we need to align with the start of the table block.
            banners_table_df_for_logic = banners_df_processed.iloc[target_idx[0][0]-2:target_idx[0][1], 1:4]
            
            # For the counts data used in the logic, we take a slice from the original counts DataFrame.
            counts_table_df = counts_df.iloc[fetch_idx[0][0]:fetch_idx[0][1], 1:4]

            # --- For WRITING the banner side (Columns B, C, D), we MUST use the ORIGINAL banner data ---
            # This ensures we are writing the exact content as it appeared in the source banners file.
            banners_table_df_for_writing = banners_df_original.iloc[target_idx[0][0]-2:target_idx[0][1], 1:4]
            
            # --- Write ORIGINAL banner data into Columns B, C, D ---
            for m, row_data in enumerate(banners_table_df_for_writing.itertuples(index=False), start=target_idx[0][0]):
                sheet.cell(row=m, column=2, value=row_data[0]) # Write label to Column B.
                sheet.cell(row=m, column=3, value=row_data[1]) # Write count to Column C.
                percent_val = row_data[2] # Get the percentage value.
                cell_d = sheet.cell(row=m, column=4) # Target cell for percentage in Column D.
                if pd.notna(percent_val):
                    # Handle potential string percentages (e.g., "50%") or numeric ones.
                    if isinstance(percent_val, str) and '%' in percent_val:
                        try:
                            numeric_percent = float(percent_val.replace('%','')) / 100 # Convert to a decimal.
                            cell_d.value = numeric_percent
                        except ValueError:
                             cell_d.value = percent_val # If conversion fails, write as is.
                    elif isinstance(percent_val, (int, float)):
                        cell_d.value = float(percent_val) # Ensure it's a float.
                    else:
                        cell_d.value = percent_val # Write any other type as is.
                    
                    # Apply the percentage format if the value is numeric.
                    if isinstance(cell_d.value, (int, float)):
                        cell_d.number_format = percentage_format
                else:
                    cell_d.value = None # If the value is missing, set the cell to None.
                    
            # --- Align Counts with Banners using the SPECIFIC Base Texts ---
            # This function takes the relevant slices of banners (for logic) and counts data,
            # and uses the `specific_base_texts_for_this_table` list to align them correctly.
            aligned_counts_df = align_counts_with_banners(banners_table_df_for_logic, counts_table_df, 'Label', specific_base_texts_for_this_table)

            if aligned_counts_df.empty: continue # If no counts data was aligned, skip to the next table.

            # --- Write the ALIGNED counts data into Columns F, G, H ---
            for k, row_data in enumerate(aligned_counts_df.itertuples(index=False), start=target_idx[0][0]):
                sheet.cell(row=k, column=6, value=row_data[0]) # Write label to Column F (Label.1).
                sheet.cell(row=k, column=7, value=row_data[1]) # Write count to Column G (Count.1).
                percent_val = row_data[2] # Get the percentage value.
                if pd.notna(percent_val):
                    cell_h = sheet.cell(row=k, column=8) # Target cell for percentage in Column H.
                    cell_h.value = float(percent_val) # Write the percentage value.
                    cell_h.number_format = percentage_format # Apply percentage formatting.
                else:
                    sheet.cell(row=k, column=8).value = None # Set to None if missing.

        except Exception as e:
            # If any error occurs during processing a table, print an error message.
            print(f"Error processing and pasting table index {i}: {e}")
            
    # After processing all tables, save the final populated workbook.
    workbook.save(file_path)
    print("Pasting of aligned table data into the output file is complete.")

def update_tab_plan_titles(file_path, counts_df, matching_df, common_indices,variablePrefix):
    """
    This function updates the table titles in the output file. It takes the refined titles
    from the `matching_df` and applies them to the correct rows in the 'Tables' sheet.
    It uses the identified table blocks to know where to make these changes.
    """
    print("Updating the Tab Plan Titles in the output file...")
    # Open the output workbook.
    workbook = load_workbook(file_path)
    sheet = workbook['Tables']
    
    # First, create a quick lookup dictionary from the `matching_df` for easy access to title information.
    label_to_info = {}
    for _, row in matching_df.iterrows():
        if pd.notna(row['Matched_Label']): # Only process rows where a match was found.
            label_to_info[row['Matched_Label']] = {
                'Title': row['Title'], # The new main title.
                'Tab Plan Titles': row['Tab Plan Titles'], # The specific title for the tab plan.
                'Original Labels': row['Original Labels'] # The original label from the banners file, used for precise targeting.
            }
            
    # Now, we iterate through each identified table block to update its titles.
    for i, (fetch_idx, target_idx) in enumerate(common_indices):
        if not fetch_idx or not target_idx: continue # Skip if indices are invalid.
        try:
            # --- 1. Identify which 'Matched_Label' this table block corresponds to ---
            # We use the first label in the count data for this block to find the related info.
            first_label_in_count_table = counts_df.iloc[fetch_idx[0][0]]['Label']
            if isinstance(first_label_in_count_table, str) and first_label_in_count_table.startswith(variablePrefix):
                variable_part = first_label_in_count_table.replace(variablePrefix, "").split(' ')[0]
                
                # Use fuzzy matching to get the precise 'Matched_Label' key. This is important if there are minor variations.
                matched_label, score = process.extractOne(variable_part, label_to_info.keys())
                if score < 100: continue # If it's not a perfect match, skip this block.

                info = label_to_info.get(matched_label) # Get all the title information for this matched label.
                if not info: continue # If no info found, skip.
                
                # --- 2. Construct the new, combined title ---
                title = info['Title']
                tab_plan_title = info['Tab Plan Titles']
                
                # We also try to grab the crosstab part (like "[AWARENESS]") from the original count file label.
                after_pipe = first_label_in_count_table.split('||')[1].strip() if '||' in first_label_in_count_table else None
                # Combine all parts to form the final title that will go into Column F.
                formatted_label_for_counts = f"{title}. [{after_pipe}] - {tab_plan_title.rstrip()}" if after_pipe else f"{title}. {tab_plan_title.rstrip()}"

                # --- 3. Find the exact row in the output file that needs the new title ---
                # We look for the row containing the *original* question text (from the banners file).
                original_question_text = info['Original Labels']
                
                start_row, end_row = target_idx[0][0], target_idx[0][1] # Get the bounds of the banner table block.
                found = False
                # Search within the relevant rows of the 'Tables' sheet. We search a bit above the start row too, just in case.
                for r in range(start_row - 2, end_row): 
                    if sheet.cell(row=r, column=2).value == original_question_text:
                        # Once we find the row with the original text, we update Column F (index 6) with our new formatted title.
                        sheet.cell(row=r, column=6).value = formatted_label_for_counts
                        
                        found = True # Mark that we've found and updated the correct row.
                        break # Stop searching for this table block.
                
                if not found:
                    # If we couldn't find the exact original label (which is rare but possible),
                    # we fall back to a slightly less precise method: updating the third row of the block.
                    # We also print a warning to alert the user.
                    print(f"Warning: Could not find exact label '{original_question_text}' for index {i}. Using fixed offset for title update.")
                    third_row_index = target_idx[0][0] + 2 # Calculate the index of the third row in the block.
                    sheet.cell(row=third_row_index, column=6).value = formatted_label_for_counts

        except Exception as e:
            # If any error occurs during title updating for a specific block, print an error message.
            print(f"Error updating tab plan title for index {i}: {e}")
            
    # Save the workbook after all titles have been updated.
    workbook.save(file_path)
    print("Tab Plan Title update process has finished.")
    
def identify_mean_tables_and_get_indices(initial_banners_df, matching_df, common_titles_map):
    """
    This function scans the original 'Banners' data to identify which tables are marked as 'Mean' tables.
    It then updates the `matching_df` to flag these tables and stores their original start/end row indices.
    This information is crucial for later statistical calculations.
    """
    matching_df['Mean_Table'] = False # Initialize a new column in `matching_df` to track mean tables.
    mean_table_indices_map = {} # This dictionary will store the original indices for identified mean tables.

    # Iterate through each table that we've successfully matched.
    for variable_name, original_banner_indices_list in common_titles_map.items():
        if not original_banner_indices_list: continue # Skip if no indices were found for this table.
        
        # Get the start and end row indices from the original banners file.
        # Remember: `target_idx[0][0]` is 1-based row index in openpyxl, so adjust for DataFrame slicing.
        original_banner_indices = original_banner_indices_list[0] # Assuming one set of indices per matched label.
        start_idx, end_idx = original_banner_indices[0] - 2, original_banner_indices[1]
        
        if start_idx < 0: start_idx = 0 # Ensure the start index is not negative.

        # Take a slice of the original banners DataFrame corresponding to this table.
        table_slice_in_original_banner = initial_banners_df.iloc[start_idx:end_idx]
        
        # Check if the label 'Mean' exists within the second column of this table slice.
        if 'Mean' in table_slice_in_original_banner.iloc[:, 1].values:
            # If 'Mean' is found, we need to find the specific row for the question label to update `matching_df`.
            # We look for rows containing a pipe '|' which often separates different parts of the banner.
            pipe_indices = table_slice_in_original_banner[table_slice_in_original_banner.iloc[:, 1] == '|'].index
            if not pipe_indices.empty:
                # Find the index of the first pipe within the slice.
                first_pipe_in_slice_idx = pipe_indices[0] - start_idx
                # Get the labels from that point onwards, ensuring we only consider rows with actual labels.
                valid_labels = table_slice_in_original_banner.iloc[first_pipe_in_slice_idx:][table_slice_in_original_banner.iloc[first_pipe_in_slice_idx:, 1].notna()]
                
                # Mean tables usually have at least 3 relevant label rows (e.g., Question, Mean, Std Dev).
                if len(valid_labels) >= 3:
                    question_label = valid_labels.iloc[2, 1] # The third label (index 2) is typically the question text.
                    
                    # Now, update the `matching_df` to flag this table as a mean table.
                    # We do this by finding the row in `matching_df` that has this specific `question_label`.
                    matching_df.loc[matching_df['Original Labels'] == question_label, 'Mean_Table'] = True
            
            # Store the original start and end indices for this mean table, keyed by its variable name.
            mean_table_indices_map[variable_name] = (start_idx, end_idx)
            
    return matching_df, mean_table_indices_map

def calculate_and_write_statistics(file_path, matching_df, mean_table_indices_map):
    """
    This function performs the statistical calculations (Mean, Standard Deviation, Standard Error)
    for the identified mean tables. It reads the data from the *current* state of the output file
    (after alignment) and writes the calculated statistics back into it.
    """
    print("Calculating Mean, Standard Deviation, and Standard Error for identified tables...")
    # Read the current state of the output file into a DataFrame.
    formatting_df = pd.read_excel(file_path, sheet_name='Tables') 
    # Open the workbook to write the results.
    workbook = load_workbook(file_path)
    sheet = workbook['Tables']
    
    # Helper function to clean numeric values, especially those that might be strings like '123'.
    def clean_numeric_value(value):
        if isinstance(value, str):
            numeric_match = re.match(r'^\d+', value) # Look for digits at the start of the string.
            if numeric_match: return int(numeric_match.group()) # Convert to integer if found.
        return value # Return the value as is if it's not a string or doesn't start with digits.

    # Helper function to extract values enclosed in parentheses at the end of a string, often used for percentages.
    def extract_parenthesis_value(label):
        if isinstance(label, str) and label.endswith(')'): # Check if the label is a string ending with ')'.
            match = re.search(r'\(([0-9.]+)\)$', label) # Find the pattern: (digits or decimal point) followed by ')'.
            if match:
                try: return float(match.group(1)) # Convert the captured number to a float.
                except ValueError: return None # Return None if conversion fails.
        return None # Return None if the pattern isn't found.

    calculated_values = [] # This list will store the calculated statistics before writing them to the file.
    formatting_df['calc_multiply'] = 1.0 # Add a temporary column for intermediate calculations.

    # Get the list of variable names for which we need to calculate statistics.
    mean_variables_to_process = matching_df[matching_df['Mean_Table'] == True]['Matched_Label'].values

    # Process each mean table.
    for var_name in mean_variables_to_process:
        if var_name not in mean_table_indices_map:
            print(f"Warning: Mean variable '{var_name}' found in matching_df but not in mean_table_indices_map. Skipping stats calculation for it.")
            continue

        start_row, end_row = mean_table_indices_map[var_name] # Get the row boundaries for this table.
        
        try:
            has_parenthesis_values = False # Flag to indicate if we found parenthesis values (for weighted calculations).
            parenthesis_values_map = {} # Store the extracted parenthesis values mapped to their row indices.

            # First pass: Check for parenthesis values and store them.
            for idx in range(start_row, end_row):
                if idx >= len(formatting_df): break # Safety check for index bounds.
                label = formatting_df.iloc[idx, 1] # Get the label from the current row.
                value = extract_parenthesis_value(label) # Try to extract a parenthesis value.
                if value is not None:
                    has_parenthesis_values = True # Set the flag if we find any.
                    parenthesis_values_map[idx] = value # Store the value and its row index.
            
            mean_val, std_dev, std_err, sigma = None, None, None, None # Initialize statistics variables.

            # Perform calculations based on whether parenthesis values were found.
            if has_parenthesis_values: 
                # If parenthesis values exist, we perform weighted calculations.
                paren_values, col_g_values = [], []
                # Reset the temporary 'calc_multiply' column for this table's range to 0, as we'll use it for weighted sums.
                formatting_df.loc[start_row:end_row-1, 'calc_multiply'] = 0.0 
                
                # Iterate through the stored parenthesis values.
                for idx, paren_value in parenthesis_values_map.items():
                    if idx >= len(formatting_df): break # Safety check.
                    # Get the corresponding value from Column G (Count.1).
                    col_g_value = pd.to_numeric(formatting_df.iloc[idx, 6], errors='coerce') 
                    if pd.notna(col_g_value):
                        paren_values.append(paren_value) # Add to list for averaging.
                        col_g_values.append(col_g_value) # Add to list for weights.
                        # Calculate the weighted product and store in the temp column.
                        formatting_df.loc[idx, 'calc_multiply'] = paren_value * col_g_value
                
                # Calculate the total sum of the weighted products.
                total = formatting_df['calc_multiply'].iloc[start_row:end_row].sum(skipna=True)
                
                # Find the 'Sigma' row in Column F to get the sample size (n).
                sigma_series = formatting_df.iloc[start_row:end_row][formatting_df.iloc[start_row:end_row, 5].astype(str).str.contains('Sigma', case=False, na=False)]['Count.1']
                if not sigma_series.empty:
                    # Convert the sigma value to numeric, handling potential errors.
                    sigma = pd.to_numeric(sigma_series.values[0], errors='coerce')
                    if pd.notna(sigma) and sigma > 0: # If sigma is valid, calculate the mean.
                        mean_val = total / sigma
                
                # Calculate weighted mean, variance, and standard deviation if we have data.
                if paren_values and col_g_values and sum(col_g_values) > 0:
                    weighted_mean = np.average(paren_values, weights=col_g_values)
                    variance = np.average((np.array(paren_values) - weighted_mean)**2, weights=col_g_values)
                    std_dev = np.sqrt(variance)
            else: 
                # If no parenthesis values, perform standard calculations on Columns F and G.
                # Clean and convert values from Columns F and G.
                data_col_f = pd.to_numeric([clean_numeric_value(str(v)) for v in formatting_df.iloc[start_row:end_row, 5].values], errors='coerce') 
                data_col_g = pd.to_numeric([clean_numeric_value(str(v)) for v in formatting_df.iloc[start_row:end_row, 6].values], errors='coerce') 
                
                # Calculate the product of Column F and G values and store in the temp column.
                formatting_df.loc[start_row:end_row-1, 'calc_multiply'] = data_col_f * data_col_g
                # Calculate the total sum of these products.
                total = formatting_df['calc_multiply'].iloc[start_row:end_row].sum(skipna=True)
                
                # Find the 'Sigma' row again to get the sample size.
                sigma_series = formatting_df.iloc[start_row:end_row][formatting_df.iloc[start_row:end_row, 5].astype(str).str.contains('Sigma', case=False, na=False)]['Count.1']
                if not sigma_series.empty:
                    sigma = pd.to_numeric(sigma_series.values[0], errors='coerce')
                    if pd.notna(sigma) and sigma > 0: mean_val = total / sigma

                # Filter out NaN values to perform calculations on valid data points.
                valid_indices = ~np.isnan(data_col_f) & ~np.isnan(data_col_g)
                values, frequencies = data_col_f[valid_indices], data_col_g[valid_indices]
                
                # Calculate weighted mean, variance, and standard deviation if we have valid data.
                if len(values) > 0 and sum(frequencies) > 0:
                    weighted_mean = np.average(values, weights=frequencies)
                    variance = np.average((values - weighted_mean)**2, weights=frequencies)
                    std_dev = np.sqrt(variance)
            
            # Calculate Standard Error if Standard Deviation and Sigma are available.
            if std_dev is not None and sigma is not None and pd.notna(sigma) and sigma > 0:
                std_err = std_dev / np.sqrt(sigma)
            
            # Now, find the specific rows for 'Mean', 'Std. Dev.', and 'Std. Err.' within the current table's data.
            table_range_df_current = formatting_df.iloc[start_row:end_row]
            
            # Find the row index for 'Mean'.
            mean_row_series = table_range_df_current[table_range_df_current.iloc[:, 1].astype(str).str.contains('Mean', case=False, na=False)].index
            if mean_val is not None and not mean_row_series.empty:
                mean_idx_in_current_df = mean_row_series[0] # Get the index within the slice.
                # Store the data to be written: row, column, and value. Add 2 to convert slice index to openpyxl row index.
                calculated_values.extend([
                    {'row': mean_idx_in_current_df + 2, 'column': 6, 'value': "Mean"}, 
                    {'row': mean_idx_in_current_df + 2, 'column': 7, 'value': round(mean_val, 2)} # Round mean to 2 decimal places.
                ])

            # Find the row index for 'Std. Dev.'.
            std_dev_row_series = table_range_df_current[table_range_df_current.iloc[:, 1].astype(str).str.contains('Std\. Dev\.', case=False, na=False)].index
            if std_dev is not None and not std_dev_row_series.empty:
                std_dev_idx_in_current_df = std_dev_row_series[0]
                calculated_values.extend([
                    {'row': std_dev_idx_in_current_df + 2, 'column': 6, 'value': "Std. Dev."},
                    {'row': std_dev_idx_in_current_df + 2, 'column': 7, 'value': round(std_dev, 2)} # Round std dev to 2 decimal places.
                ])

            # Find the row index for 'Std. Err.'.
            std_err_row_series = table_range_df_current[table_range_df_current.iloc[:, 1].astype(str).str.contains('Std\. Err\.', case=False, na=False)].index
            if std_err is not None and not std_err_row_series.empty:
                std_err_idx_in_current_df = std_err_row_series[0]
                calculated_values.extend([
                    {'row': std_err_idx_in_current_df + 2, 'column': 6, 'value': "Std. Err."},
                    {'row': std_err_idx_in_current_df + 2, 'column': 7, 'value': round(std_err, 2)} # Round std err to 2 decimal places.
                ])
        except Exception as e:
            # If an error occurs during calculation for a specific table, print a message.
            print(f"An error occurred calculating stats for var_name '{var_name}' in table range {start_row}-{end_row}: {e}")

    # If any statistics were calculated, write them back to the Excel file.
    if calculated_values:
        for item in calculated_values:
            try:
                # Write the calculated value to the correct cell.
                sheet.cell(row=item['row'], column=item['column']).value = item['value']
            except Exception as cell_e:
                # Handle potential errors during cell writing.
                print(f"Error writing to cell (Row: {item['row']}, Col: {item['column']}): {cell_e}")
        workbook.save(file_path) # Save the workbook with the new statistics.
        print("Statistics calculation and writing complete.")

def calculate_box_summaries(file_path):
    """
    This function calculates the 'Top Box' and 'Bottom Box' summaries. It looks for specific
    labels like 'Top X Box (Net)' and sums the corresponding percentage values below them.
    """
    print("Calculating Top/Bottom Box values...")
    # Read the current state of the output file.
    formatting_df = pd.read_excel(file_path, sheet_name='Tables')
    # Open the workbook to write the results.
    workbook = load_workbook(file_path)
    sheet = workbook['Tables']
    
    # Iterate through each row of the DataFrame.
    for idx, row in formatting_df.iterrows():
        # Use a regular expression to find labels like 'Top 2 Box (Net)' or 'Bottom 3 Box (Net)'.
        match = re.match(r"(Top|Bottom) (\d+) Box \(Net\)", str(row['Label']), re.IGNORECASE)
        if match:
            direction, num_rows_str = match.groups() # Extract 'Top'/'Bottom' and the number of rows.
            num_rows = int(num_rows_str) # Convert the number of rows to an integer.
            
            # Sum the values from Column G (Count.1) for the specified number of rows immediately following the 'Net' label.
            # We convert values to numeric, coercing errors to NaN.
            values_to_sum = pd.to_numeric(formatting_df.iloc[idx + 1: idx + 1 + num_rows, 6], errors='coerce')
            box_sum = values_to_sum.sum() # Calculate the sum.
            
            # Write the calculated summary value back into the correct row and columns in the Excel sheet.
            # We update the label in Column F and the sum in Column G.
            sheet.cell(row=idx + 2, column=6).value = f"{direction} {num_rows} Box (Net)" # Correctly format the label in Col F.
            sheet.cell(row=idx + 2, column=7).value = box_sum # Write the calculated sum in Col G.
            
    workbook.save(file_path) # Save the workbook with the updated box summaries.
    print("Top/Bottom Box summaries update complete.")

def add_comparison_and_formatting(file_path):
    """
    Adding comparison columns and applying conditional formatting to highlight matches/mismatches.
    This function adds the final comparison columns (J, K, L) with formulas and applies
    conditional formatting to highlight matches and mismatches between the banner data
    (Columns B, C, D) and the aligned count data (Columns F, G, H).
    """
    print("Adding final comparison columns and applying conditional formatting...")
    # Open the workbook.
    workbook = load_workbook(file_path)
    sheet = workbook['Tables']
    
    # We need to read the *values* from the sheet as they are *before* adding formulas,
    # so we can use these values to set the initial fills correctly based on current matches.
    # Loading with `data_only=True` ensures we get the displayed values (or results of existing formulas).
    temp_wb_for_values = load_workbook(file_path, data_only=True)
    temp_ws_for_values = temp_wb_for_values['Tables']

    max_row = temp_ws_for_values.max_row # Get the maximum row number from the data-loaded sheet.
    
    # Define the fill colors for matches (greenish) and mismatches (reddish).
    red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE") # Light red for mismatch.
    purple_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE") # Light green for match.

    # Iterate through each row to add formulas and apply initial fills.
    for r in range(1, max_row + 1):
        # --- Add Comparison Formulas ---
        # Column J: Compares the original label (B) with the aligned count label (F).
        sheet.cell(row=r, column=10).value = f'=IF(EXACT(B{r},F{r}),"TRUE","FALSE")'
        # Column K: Compares the original count (C) with the aligned count count (G).
        sheet.cell(row=r, column=11).value = f'=IF(C{r}=G{r},"TRUE","FALSE")'
        # Column L: Compares the original percentage (D) with the aligned count percentage (H).
        sheet.cell(row=r, column=12).value = f'=IF(D{r}=H{r},"TRUE","FALSE")'
        
        # --- Apply Initial Fills based on current data ---
        # Get values from the temp sheet to check for initial matches.
        val_b = temp_ws_for_values.cell(row=r, column=2).value
        val_f = temp_ws_for_values.cell(row=r, column=6).value
        # Apply fill to Column J based on the comparison of B and F. Use string comparison for labels.
        sheet.cell(row=r, column=10).fill = purple_fill if str(val_b) == str(val_f) else red_fill
        
        val_c = temp_ws_for_values.cell(row=r, column=3).value
        val_g = temp_ws_for_values.cell(row=r, column=7).value
        try:
            # Compare counts C and G. Use a small tolerance for float comparisons to handle potential precision issues.
            is_equal_cg = (pd.isna(val_c) and pd.isna(val_g)) or (abs(float(val_c) - float(val_g)) < 1e-9) 
        except (ValueError, TypeError):
            is_equal_cg = str(val_c) == str(val_g) # Fallback to string comparison if values are not numeric.
        # Apply fill to Column K based on the comparison of C and G.
        sheet.cell(row=r, column=11).fill = purple_fill if is_equal_cg else red_fill

        val_d = temp_ws_for_values.cell(row=r, column=4).value # Original percentage (should be numeric).
        val_h = temp_ws_for_values.cell(row=r, column=8).value # Aligned percentage (should be numeric).
        try:
            # Compare percentages D and H using a tolerance for floating-point accuracy.
            is_equal_dh = (pd.isna(val_d) and pd.isna(val_h)) or \
                          (pd.notna(val_d) and pd.notna(val_h) and abs(float(val_d) - float(val_h)) < 1e-9) 
        except (ValueError, TypeError):
             is_equal_dh = str(val_d) == str(val_h) # Fallback if values are not numbers.
        # Apply fill to Column L based on the comparison of D and H.
        sheet.cell(row=r, column=12).fill = purple_fill if is_equal_dh else red_fill

    # --- Apply Conditional Formatting Rules ---
    # These rules ensure that the colors update automatically if the underlying formula results change.
    # For Column J: if the formula result is "TRUE", apply purple fill; if "FALSE", apply red fill.
    sheet.conditional_formatting.add(f'J1:J{max_row}', FormulaRule(formula=[f'J1="TRUE"'], fill=purple_fill))
    sheet.conditional_formatting.add(f'J1:J{max_row}', FormulaRule(formula=[f'J1="FALSE"'], fill=red_fill))
    # Repeat for Column K.
    sheet.conditional_formatting.add(f'K1:K{max_row}', FormulaRule(formula=[f'K1="TRUE"'], fill=purple_fill))
    sheet.conditional_formatting.add(f'K1:K{max_row}', FormulaRule(formula=[f'K1="FALSE"'], fill=red_fill))
    # Repeat for Column L.
    sheet.conditional_formatting.add(f'L1:L{max_row}', FormulaRule(formula=[f'L1="TRUE"'], fill=purple_fill))
    sheet.conditional_formatting.add(f'L1:L{max_row}', FormulaRule(formula=[f'L1="FALSE"'], fill=red_fill))

    workbook.save(file_path) # Save the workbook with formulas and conditional formatting applied.
    print("Final comparison columns and conditional formatting have been added.")
    
    
    # --- VALIDATOR FUNCTIONS ---
# These functions run pre-flight checks to ensure the script has everything it needs to run successfully.
# They check for dependencies, paths, files, and basic data integrity.

def validate_dependencies():
    """Checks if all required third-party libraries are installed."""
    print("Validator: Checking for required libraries...")
    try:
        import pandas
        import numpy
        import fuzzywuzzy
        import openpyxl
        # 're', 'os', 'collections', 'copy' are standard libraries, no need to check.
    except ImportError as e:
        library_name = str(e).split("'")[1]
        print(f"ERROR: Missing required library '{library_name}'.")
        print(f"Please install it by running: pip install {library_name}")
        raise e
    print(" -> Dependencies are satisfied.")

def validate_paths_and_files(input_folder, output_folder, counts_file, banners_file, matching_file,CountFileName,BannerFileName,MatchFileName):
    """
    Validates the existence of input/output folders and the required source files.
    Also checks if the output folder is writable.
    """
    print("Validator: Checking folder paths and file existence...")
    # 1. Check if input folder exists
    if not os.path.isdir(input_folder):
        raise FileNotFoundError(f"Input folder not found at: {input_folder}")

    # 2. Check if output folder exists
    if not os.path.isdir(output_folder):
        raise FileNotFoundError(f"Output folder not found at: {output_folder}")

    # 3. Check for write permissions in the output folder
    try:
        test_file = os.path.join(output_folder, 'permission_test.tmp')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
    except (IOError, OSError) as e:
        raise PermissionError(f"Cannot write to the output folder: {output_folder}. Please check permissions. Error: {e}")

    # 4. Check if all required input files exist
    for file_path, file_desc in [
        (counts_file, f"{CountFileName}.xlsx"),
        (banners_file, f"{BannerFileName}.xlsx"),
        (matching_file, f"{MatchFileName}.xlsx")
    ]:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Required input file not found: {file_desc} at {file_path}")
            
    print(" -> Paths and files are valid.")

def validate_excel_schema(file_path, required_sheet, required_columns, file_description):
    """
    A generic validator to check for a specific sheet and column headers in an Excel file.
    """
    print(f"Validator: Checking schema for '{file_description}'...")
    # Check for the required sheet
    xls = pd.ExcelFile(file_path)
    if required_sheet not in xls.sheet_names:
        raise ValueError(f"File '{file_description}' is missing the required sheet: '{required_sheet}'. Found sheets: {xls.sheet_names}")
    
    # Check for required columns without loading the whole file
    df_cols = pd.read_excel(file_path, sheet_name=required_sheet, nrows=0).columns
    missing_cols = [col for col in required_columns if col not in df_cols]
    
    if missing_cols:
        raise ValueError(f"File '{file_description}' (sheet: '{required_sheet}') is missing required columns: {missing_cols}. Found columns: {list(df_cols)}")

    print(f" -> Schema for '{file_description}' is valid.")

def validate_counts_file_content(counts_path, prefix,CountFileName):
    """
    Checks the Counts.xlsx file for the presence of the essential VARIABLE_PREFIX.
    If this prefix is missing, the script cannot identify any tables to process.
    """
    print(f"Validator: Checking content integrity of '{CountFileName}.xlsx'...")
    df = pd.read_excel(counts_path, sheet_name='Tables', header=None)
    
    # Check if any cell in the first column contains the prefix
    if not df[0].astype(str).str.contains(prefix, regex=False, na=False).any():
        raise ValueError(f"Data Integrity Error in '{CountFileName}.xlsx': No rows containing the prefix '{prefix}' were found. The script cannot identify table variables without this marker.")
    print(f" -> Content integrity for '{CountFileName}.xlsx' is OK.")

def run_pre_flight_checks(InputDir,OutputDir,CountFileName,BannerFileName,MatchFileName,variablePrefix):
    """
    Main validator function to run all checks in sequence.
    If any check fails, it will raise an exception and stop the script.
    """
    print("--- Running Pre-flight Validation Checks ---")
    
    # 1. Check for library dependencies
    validate_dependencies()
    
    # 2. Define file paths for validation
    counts_path = os.path.join(InputDir, f'{CountFileName}.xlsx')
    banners_path = os.path.join(InputDir, f'{BannerFileName}.xlsx')
    matching_path = os.path.join(OutputDir, f'{MatchFileName}.xlsx')
    
    # 3. Check paths and file existence
    validate_paths_and_files(InputDir, OutputDir, counts_path, banners_path, matching_path,CountFileName,BannerFileName,MatchFileName)
    
    # 4. Check schema of Banners.xlsx
    # We don't specify columns as the script uses iloc, but we must check for the sheet.
    validate_excel_schema(banners_path, 'Tables', [], f"{BannerFileName}.xlsx")

    # 5. Check schema of Counts.xlsx
    # The script renames columns, so we just check that there are at least 4.
    counts_cols = pd.read_excel(counts_path, sheet_name='Tables').columns
    if len(counts_cols) != 4:
         raise ValueError(f"File '{CountFileName}.xlsx' must have 4 columns. Found {len(counts_cols)}.")
    
    # 6. Check schema of Matched_Variables.xlsx
    required_matching_cols = ['Matched_Label', 'Title', 'Tab Plan Titles', 'Original Labels', 'Base Text']
    validate_excel_schema(matching_path, 'Sheet1', required_matching_cols, f"{MatchFileName}.xlsx")
    
    # 7. Check critical content of Counts.xlsx
    validate_counts_file_content(counts_path, variablePrefix,CountFileName)
    
    print("--- All validation checks passed successfully. ---\n")

def main(InputDir,OutputDir,CountFileName,BannerFileName,MatchFileName,FinalComparisonFileName,SummaryFileName,variablePrefix):
    """
    This is the main orchestrator function. It calls all the other functions in the correct
    sequence to perform the entire data comparison and formatting process.
    """
    try:
        # --- 0. RUN VALIDATORS ---
        # This will stop the script if any prerequisite is not met.
        run_pre_flight_checks(InputDir,OutputDir,CountFileName,BannerFileName,MatchFileName,variablePrefix)

        # --- 1. SETUP: Define file paths and load initial data ---
        print("Starting the main process...")
        # Construct the full paths to the input and output files.
        counts_path = os.path.join(InputDir, f'{CountFileName}.xlsx')
        banners_path = os.path.join(InputDir, f'{BannerFileName}.xlsx')
        matching_path = os.path.join(OutputDir, f'{MatchFileName}.xlsx')
        final_comparison_path = os.path.join(OutputDir, f'{FinalComparisonFileName}.xlsx')
        summary_path = os.path.join(OutputDir, f'{SummaryFileName}.txt')

        print("Loading source files (Banners, Counts, Matching Results)...")
        # Load the initial Banners data, which serves as our structural template.
        banners_df_initial = pd.read_excel(banners_path, sheet_name='Tables')
        # Load the Counts data.
        counts_df = pd.read_excel(counts_path, sheet_name='Tables')
        # Rename the columns of the Counts DataFrame for clarity.
        counts_df.columns = ['Names', 'Label', 'Counts', 'Percentile']
        # Load the results from the matching process.
        matching_df = pd.read_excel(matching_path)

        # --- PREPARE BASE TEXTS MAP ---
        # This map is crucial for correctly associating specific 'Base Text' entries with their corresponding tables.
        # It's built from the `matching_df` before any processing begins.
        base_texts_map = defaultdict(list)
        for _, row in matching_df.iterrows():
            if pd.notna(row['Matched_Label']): # Only consider rows where a match was found.
                base_texts_map[row['Matched_Label']].append(row['Base Text']) # Store the base text associated with the matched label.

        # --- 2. GENERATE UNMATCHED SUMMARY ---
        # Create a summary of tables that couldn't be matched automatically.
        # We pass a copy of matching_df to avoid modifying the original DataFrame passed to other functions.
        generate_unmatched_summary(matching_df.copy(), summary_path)

        # --- 3. PREPARE OUTPUT FILE & IDENTIFY TABLE STRUCTURE ---
        print("Preparing the output file structure based on the Banners file...")
        # Create the main output file and copy the formatting from the Banners file.
        create_and_prepare_output_file(banners_path, final_comparison_path)
        
        # Create a mapping from original banner labels to their matched labels. This is used to process the banner data correctly.
        match_dict = defaultdict(list)
        for original, matched_label in zip(matching_df['Original Labels'], matching_df['Matched_Label']):
            match_dict[original].append(matched_label)
        
        # Create a processed version of the initial banners DataFrame. This step is a bit complex;
        # it seems to be about ensuring that labels in the banners file are correctly mapped or transformed
        # before we use them to find corresponding tables in the counts file.
        banners_df_processed = banners_df_initial.copy() 
        # This line attempts to replace labels in the processed banner DataFrame with their matched counterparts if available.
        banners_df_processed.iloc[:, 1] = banners_df_processed.iloc[:, 1].apply(lambda v: match_dict.get(v, [v]).pop(0) if match_dict.get(v) else v)
        
        # Extract all unique variable names from the Counts file.
        count_variable_names = get_variable_names_from_counts(counts_df, variablePrefix)
        # Get all the unique 'Matched_Labels' from our matching results. These are the targets we're looking for.
        banner_titles_from_matching = matching_df['Matched_Label'].dropna().unique()
        
        # This dictionary will store mappings from matched banner labels to their corresponding indices in the counts file.
        common_titles_map = {} 
        # This list will hold the final index pairs (fetch_idx, target_idx) needed for populating the data.
        expanded_indices_for_population = [] 

        # Now, we iterate through each potential table title (from the matching results)
        # and try to find its corresponding location in both the Counts and Banners files.
        for matched_label in banner_titles_from_matching:
            # Find the best matching variable name in the Counts file for the current matched banner label.
            # We use `token_sort_ratio` for a more flexible matching.
            best_match_in_counts_tuple = process.extractOne(matched_label, count_variable_names, scorer=fuzz.token_sort_ratio)
            if not best_match_in_counts_tuple : continue # If no match is found, skip this banner label.
            
            best_match_in_counts, score = best_match_in_counts_tuple
            # We only proceed if the match score is high enough (95 or above).
            if score >= 95:
                # Find the row indices for this matched variable name in the Counts file.
                fetch_indices = fetch_tables_indices(counts_df, variablePrefix, best_match_in_counts, 'Count') 
                # Find the row indices for this matched label in the processed Banners file.
                target_indices = fetch_tables_indices(banners_df_processed, variablePrefix, matched_label, 'Banner')
                
                # If we successfully found indices in both files, we store them.
                if fetch_indices and target_indices:
                    common_titles_map[matched_label] = target_indices # Store the banner indices for later use (e.g., stats).

                    # The `fetch_tables_indices` function can return multiple index pairs for a single table title
                    # if there are multiple instances or complex structures. We need to "expand" these so that
                    # each individual table block gets processed.
                    if len(target_indices) > 1:
                        for sub_tuple_target in target_indices:
                            # For each sub-tuple of target indices, create a corresponding entry for population.
                            expanded_indices_for_population.append((fetch_indices, [sub_tuple_target]))
                    else:
                        # If there's only one set of indices, add it directly.
                        expanded_indices_for_population.append((fetch_indices, target_indices))
        
        # --- IDENTIFY MEAN TABLES ---
        # Use the identified common tables and the original banners data to flag which tables need mean calculations
        # and to get their original row indices for the calculation step.
        matching_df, mean_table_original_indices_map = identify_mean_tables_and_get_indices(banners_df_initial, matching_df.copy(), common_titles_map)
                
        # --- 4. POPULATE DATA INTO OUTPUT FILE ---
        # Now, take the aligned data and populate it into the prepared output file.
        # We pass the `base_texts_map` here to ensure correct alignment of base values.
        populate_comparison_sheet(final_comparison_path, counts_df, banners_df_processed, banners_df_initial, expanded_indices_for_population, base_texts_map,variablePrefix)
        
        # --- 5. UPDATE TABLE PLAN TITLES ---
        # Update the titles in the output file based on the matching results.
        update_tab_plan_titles(final_comparison_path, counts_df, matching_df, expanded_indices_for_population,variablePrefix)
        
        # --- 6. CALCULATE STATISTICS ---
        # Calculate and write Mean, Std Dev, and Std Err for the identified mean tables.
        calculate_and_write_statistics(final_comparison_path, matching_df, mean_table_original_indices_map)
        
        # --- 7. CALCULATE BOX SUMMARIES ---
        # Calculate and write the Top/Bottom Box summaries.
        calculate_box_summaries(final_comparison_path)
        
        # --- 8. ADD FINAL COMPARISON AND FORMATTING ---
        # Add comparison columns and apply conditional formatting to highlight matches/mismatches.
        add_comparison_and_formatting(final_comparison_path)

        print("\nScript has finished executing. All tasks completed successfully!")
        
    except (FileNotFoundError, ValueError, PermissionError, ImportError) as e:
        print(f"\nVALIDATION FAILED: A critical error occurred that prevented the script from running.")
        print(f"Error Details: {e}")
        sys.exit(1) # Exit the script with an error code
    except Exception as e:
        print(f"\nAn unexpected error occurred during execution: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1) # Exit the script with an error code
# This ensures that the main() function is called only when the script is executed directly.
if __name__ == "__main__":
    main()